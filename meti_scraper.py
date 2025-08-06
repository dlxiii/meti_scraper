from datetime import datetime, timedelta
from pathlib import Path

import csv
from openpyxl import load_workbook
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import urllib3
import pdfplumber
import pytesseract


class meti:
    """Downloader for METI data such as LNG stock and industrial production."""

    _URL = (
        "https://www.enecho.meti.go.jp/category/electricity_and_gas/"
        "electricity_measures/pdf/denryoku_LNG_stock.pdf"
    )

    def lng_weekly_inventory(self, date: str) -> str:
        """Download weekly LNG inventory PDF for the given date.

        Parameters
        ----------
        date: str
            Date string in YYYYMMDD format used in the filename.

        Returns
        -------
        str
            Path to the downloaded PDF file.
        """
        directory = Path("pdf") / "lng" / "stock"
        directory.mkdir(parents=True, exist_ok=True)
        file_path = directory / f"denryoku_LNG_stock_{date}.pdf"

        session = requests.Session()
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        retry = Retry(
            total=5,
            backoff_factor=1,
            status_forcelist=[500, 502, 503, 504],
            allowed_methods=["GET"],
        )
        adapter = HTTPAdapter(max_retries=retry)
        session.mount("https://", adapter)
        session.mount("http://", adapter)

        try:
            response = session.get(
                self._URL,
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=10,
                verify=False,
            )
            response.raise_for_status()
        except requests.exceptions.RequestException as err:
            raise RuntimeError("Failed to download METI LNG stock PDF") from err

        file_path.write_bytes(response.content)
        print(file_path)

        return str(file_path)

    def index_of_industrial_production(self, date: str) -> list[str]:
        """Download industrial production index files.

        This method retrieves three datasets related to the index of
        industrial production and saves them with Japanese filenames.
        For the dataset titled "過去の製造工業生産能力・稼働率指数（接続指数）",
        the sheet "生産付加価値額" is exported as a CSV file. For the dataset
        titled "生産・出荷・在庫・在庫率指数", the sheet named "生産" is
        additionally exported as a CSV file. For the dataset titled
        "過去の生産・出荷・在庫・在庫率指数（接続指数），鉱工業総合のみ"
        "（暦年・年度・四半期）（1953年～）", the sheets "四半期（季調）" and
        "年度（原）" are also exported as CSV files.

        Parameters
        ----------
        date: str
            Date string in YYYYMM format appended to each filename.

        Returns
        -------
        list[str]
            Paths to the downloaded Excel files and generated CSV file.
        """

        sources = {
            "過去の製造工業生産能力・稼働率指数（接続指数）":
                "https://www.meti.go.jp/statistics/tyo/iip/xls/b2020_sgs1j.xlsx",
            "生産・出荷・在庫・在庫率指数":
                "https://www.meti.go.jp/statistics/tyo/iip/xls/b2020_gsm1j.xlsx",
            "過去の生産・出荷・在庫・在庫率指数（接続指数），鉱工業総合のみ（暦年・年度・四半期）（1953年～）":
                "https://www.meti.go.jp/statistics/tyo/iip/xls/b2020_sosq1j.xlsx",
        }

        directory = Path("xls") / "iip"
        directory.mkdir(parents=True, exist_ok=True)

        session = requests.Session()
        retry = Retry(
            total=5,
            backoff_factor=1,
            status_forcelist=[500, 502, 503, 504],
            allowed_methods=["GET"],
        )
        adapter = HTTPAdapter(max_retries=retry)
        session.mount("https://", adapter)
        session.mount("http://", adapter)

        file_paths: list[str] = []
        headers = {"User-Agent": "Mozilla/5.0"}
        for name, url in sources.items():
            if name == "過去の製造工業生産能力・稼働率指数（接続指数）":
                sanitized_name = "過去の製造工業生産能力_稼働率指数_接続指数"
            elif (
                name
                == "過去の生産・出荷・在庫・在庫率指数（接続指数），鉱工業総合のみ（暦年・年度・四半期）（1953年～）"
            ):
                sanitized_name = (
                    "過去の生産_出荷_在庫_在庫率指数_接続指数_"
                    "鉱工業総合のみ_暦年_年度_四半期_1953年"
                )
            else:
                sanitized_name = name.replace("・", "_")
            file_path = directory / f"{sanitized_name}_{date}.xlsx"
            try:
                response = session.get(url, headers=headers, timeout=10)
                response.raise_for_status()
            except requests.exceptions.RequestException as err:
                raise RuntimeError(f"Failed to download {name}") from err
            file_path.write_bytes(response.content)
            file_paths.append(str(file_path))

            if name == "生産・出荷・在庫・在庫率指数":
                wb = load_workbook(file_path, data_only=True, read_only=True)
                if "生産" in wb.sheetnames:
                    ws = wb["生産"]
                    csv_path = directory / f"{sanitized_name}_生産_{date}.csv"
                    with csv_path.open("w", newline="", encoding="utf-8") as csvfile:
                        writer = csv.writer(csvfile)
                        for row in ws.iter_rows(values_only=True):
                            writer.writerow(
                                [cell if cell is not None else "" for cell in row]
                            )
                    file_paths.append(str(csv_path))
                wb.close()
            elif name == "過去の製造工業生産能力・稼働率指数（接続指数）":
                wb = load_workbook(file_path, data_only=True, read_only=True)
                if "生産付加価値額" in wb.sheetnames:
                    ws = wb["生産付加価値額"]
                    csv_path = directory / (
                        f"{sanitized_name}_生産付加価値額_{date}.csv"
                    )
                    with csv_path.open("w", newline="", encoding="utf-8") as csvfile:
                        writer = csv.writer(csvfile)
                        for row in ws.iter_rows(values_only=True):
                            writer.writerow(
                                [cell if cell is not None else "" for cell in row]
                            )
                    file_paths.append(str(csv_path))
                wb.close()
            elif (
                name
                == "過去の生産・出荷・在庫・在庫率指数（接続指数），鉱工業総合のみ（暦年・年度・四半期）（1953年～）"
            ):
                wb = load_workbook(file_path, data_only=True, read_only=True)
                sheet_map = {
                    "四半期（季調）": "四半期_季調",
                    "年度（原）": "年度_原",
                }
                for sheet, suffix in sheet_map.items():
                    if sheet in wb.sheetnames:
                        ws = wb[sheet]
                        csv_path = directory / f"{sanitized_name}_{suffix}_{date}.csv"
                        with csv_path.open("w", newline="", encoding="utf-8") as csvfile:
                            writer = csv.writer(csvfile)
                            for row in ws.iter_rows(values_only=True):
                                writer.writerow(
                                    [cell if cell is not None else "" for cell in row]
                                )
                        file_paths.append(str(csv_path))
                wb.close()

        return file_paths

    def index_of_tertiary_industry_activity(self, date: str) -> list[str]:
        """Download tertiary industry activity index files.

        This method retrieves two datasets related to the index of
        tertiary industry activity and saves them with Japanese filenames.
        The datasets are provided by METI at the following URLs:

        - ``https://www.meti.go.jp/statistics/tyo/sanzi/result/excel/b2020_ksij.xlsx``
        - ``https://www.meti.go.jp/statistics/tyo/sanzi/result/excel/b2020_ITA_linkj.xlsx``

        The workbook titled ``第３次産業活動指数_季節指数`` contains only one
        sheet, which is exported as a CSV file. For the workbook titled
        ``第３次産業活動指数_接続指数``, the sheet ``季調済指数`` is exported
        as a CSV file.

        Parameters
        ----------
        date: str
            Date string in YYYYMM format appended to each filename.

        Returns
        -------
        list[str]
            Paths to the downloaded Excel files and generated CSV file.
        """

        directory = Path("xls") / "ita"
        directory.mkdir(parents=True, exist_ok=True)

        session = requests.Session()
        retry = Retry(
            total=5,
            backoff_factor=1,
            status_forcelist=[500, 502, 503, 504],
            allowed_methods=["GET"],
        )
        adapter = HTTPAdapter(max_retries=retry)
        session.mount("https://", adapter)
        session.mount("http://", adapter)

        headers = {"User-Agent": "Mozilla/5.0"}
        file_paths: list[str] = []

        seasonal_name = f"第３次産業活動指数_季節指数_{date}.xlsx"
        seasonal_url = (
            "https://www.meti.go.jp/statistics/tyo/sanzi/result/excel/b2020_ksij.xlsx"
        )
        try:
            response = session.get(seasonal_url, headers=headers, timeout=10)
            response.raise_for_status()
        except requests.exceptions.RequestException as err:
            raise RuntimeError("Failed to download seasonal index") from err
        seasonal_path = directory / seasonal_name
        seasonal_path.write_bytes(response.content)
        file_paths.append(str(seasonal_path))

        wb = load_workbook(seasonal_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        seasonal_csv = directory / f"第３次産業活動指数_季節指数_{date}.csv"
        with seasonal_csv.open("w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([cell if cell is not None else "" for cell in row])
        wb.close()
        file_paths.append(str(seasonal_csv))

        connected_name = f"第３次産業活動指数_接続指数_{date}.xlsx"
        connected_url = (
            "https://www.meti.go.jp/statistics/tyo/sanzi/result/excel/b2020_ITA_linkj.xlsx"
        )
        try:
            response = session.get(connected_url, headers=headers, timeout=10)
            response.raise_for_status()
        except requests.exceptions.RequestException as err:
            raise RuntimeError("Failed to download connected index") from err
        connected_path = directory / connected_name
        connected_path.write_bytes(response.content)
        file_paths.append(str(connected_path))

        wb = load_workbook(connected_path, data_only=True, read_only=True)
        if "季調済指数" in wb.sheetnames:
            ws = wb["季調済指数"]
            connected_csv = directory / f"第３次産業活動指数_接続指数_季調済指数_{date}.csv"
            with connected_csv.open("w", newline="", encoding="utf-8") as csvfile:
                writer = csv.writer(csvfile)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow([cell if cell is not None else "" for cell in row])
            file_paths.append(str(connected_csv))
        wb.close()

        return file_paths

    def _extract_text(self, page) -> str:
        """Extract text from a PDF page, falling back to OCR if needed."""
        text = page.extract_text(layout=True)
        if text:
            lines = [line.strip() for line in text.splitlines()]
            while lines and not lines[0]:
                lines.pop(0)
            while lines and not lines[-1]:
                lines.pop()
            return "\n".join(lines)
        try:
            image = page.to_image(resolution=300).original
            ocr_text = pytesseract.image_to_string(image)
            lines = [line.strip() for line in ocr_text.splitlines()]
            while lines and not lines[0]:
                lines.pop(0)
            while lines and not lines[-1]:
                lines.pop()
            return "\n".join(lines)
        except Exception:
            return ""

    def _table_to_markdown(self, table) -> str:
        """Convert a table (list of lists) to a Markdown string."""
        if not table:
            return ""
        header = [cell or "" for cell in table[0]]
        md_lines = ["| " + " | ".join(header) + " |"]
        md_lines.append("| " + " | ".join("---" for _ in header) + " |")
        for row in table[1:]:
            md_lines.append("| " + " | ".join(cell or "" for cell in row) + " |")
        return "\n".join(md_lines)

    def _table_to_csv(self, table, csv_path: Path) -> None:
        """Write a table (list of lists) to a CSV file."""
        with csv_path.open("w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            for row in table:
                # Skip completely empty rows to avoid blank lines in the output
                if not any(
                    cell is not None and str(cell).strip() for cell in row
                ):
                    continue
                writer.writerow([cell or "" for cell in row])

    def pdf_tables_to_csv(self, pdf_path: str) -> list[str]:
        """Extract tables from a PDF and save them as CSV files.

        If multiple tables are found, an index is appended to the base filename.

        Parameters
        ----------
        pdf_path : str
            Path to the source PDF file.

        Returns
        -------
        list[str]
            Paths to the generated CSV files.
        """

        pdf_file = Path(pdf_path)
        tables = []
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                tables.extend(page.extract_tables() or [])

        csv_paths: list[str] = []
        if not tables:
            return csv_paths

        if len(tables) == 1:
            csv_path = pdf_file.with_suffix(".csv")
            self._table_to_csv(tables[0], csv_path)
            csv_paths.append(str(csv_path))
        else:
            for idx, table in enumerate(tables, 1):
                csv_path = pdf_file.with_name(
                    f"{pdf_file.stem}_table_{idx:03d}.csv"
                )
                self._table_to_csv(table, csv_path)
                csv_paths.append(str(csv_path))

        return csv_paths

    def pdf_to_markdown(self, pdf_path: str, include_images: bool = True) -> str:
        """Convert a PDF to Markdown, preserving tables when possible.

        This method extracts text and tables from each page. If a page
        lacks embedded text, OCR is used as a fallback.  Optionally the
        images found in the PDF can be exported and referenced in the
        generated Markdown.

        Parameters
        ----------
        pdf_path : str
            Path to the source PDF file.
        include_images : bool, default ``True``
            If ``True`` images found in the PDF are saved to PNG files and
            referenced in the Markdown.  If ``False`` images are ignored.

        Returns
        -------
        str
            Path to the generated Markdown file.
        """

        pdf_file = Path(pdf_path)
        md_file = pdf_file.with_suffix(".md")

        parts = []
        image_counter = 1
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                text = self._extract_text(page)
                if text:
                    parts.append(text)
                for table in page.extract_tables() or []:
                    md_table = self._table_to_markdown(table)
                    if md_table:
                        parts.append(md_table)
                if include_images:
                    for img in page.images:
                        bbox = (
                            img["x0"],
                            page.height - img["y1"],
                            img["x1"],
                            page.height - img["y0"],
                        )
                        cropped = page.crop(bbox)
                        pil_img = cropped.to_image(resolution=300).original
                        image_path = pdf_file.with_name(
                            f"{pdf_file.stem}_{image_counter:03d}.png"
                        )
                        pil_img.save(image_path)
                        parts.append(
                            f"![Figure {image_counter:03d}]({image_path.name})"
                        )
                        image_counter += 1

        md_file.write_text("\n\n".join(parts), encoding="utf-8")
        return str(md_file)

    def pdf_to_markdown_plain(self, pdf_path: str) -> str:
        """Convert a PDF to Markdown with text extraction only."""

        pdf_file = Path(pdf_path)
        md_file = pdf_file.with_suffix(".md")

        text_parts = []
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                text = self._extract_text(page)
                if text:
                    text_parts.append(text)

        md_file.write_text("\n\n".join(text_parts), encoding="utf-8")
        return str(md_file)


if __name__ == "__main__":
    today = datetime.today()
    weekday = today.weekday()  # Monday = 0, Sunday = 6

    if weekday >= 2:
        monday = today - timedelta(days=weekday)
        target_wed = monday + timedelta(days=2)
    else:
        monday = today - timedelta(days=weekday + 7)
        target_wed = monday + timedelta(days=2)

    scraper = meti()

    ita_paths = scraper.index_of_tertiary_industry_activity(date=today.strftime("%Y%m"))
    for path in ita_paths:
        print(path)

    iip_paths = scraper.index_of_industrial_production(date=today.strftime("%Y%m"))
    for path in iip_paths:
        print(path)

    pdf_path = scraper.lng_weekly_inventory(date=target_wed.strftime("%Y%m%d"))
    md_path = scraper.pdf_to_markdown(pdf_path)
    print(md_path)
    csv_paths = scraper.pdf_tables_to_csv(pdf_path)
    for path in csv_paths:
        print(path)
