from pathlib import Path
from urllib.parse import urljoin

import csv
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from meti_scraper import meti


class esri:
    """Downloader for ESRI GDP data."""

    _BASE_URL = "https://www.esri.cao.go.jp"
    _PAGE = _BASE_URL + "/jp/sna/menu.html"
    _KP23_PAGE = _BASE_URL + "/jp/sna/sonota/kotei/kotei_top.html"

    def gdp(self, date: str) -> list[str]:
        """Download GDP CSV files for the given date.

        Parameters
        ----------
        date: str
            Date string in YYYYMM format appended to each filename.

        Returns
        -------
        list[str]
            Paths to the downloaded CSV files.
        """

        directory = Path("csv") / "gdp"
        directory.mkdir(parents=True, exist_ok=True)

        session = requests.Session()
        retry = Retry(
            total=5,
            backoff_factor=1,
            status_forcelist=[500, 502, 503, 504],
            allowed_methods=["GET"],
        )
        adapter = HTTPAdapter(max_retries=retry)
        session.mount("https://", adapter)
        session.mount("http://", adapter)

        try:
            response = session.get(
                self._PAGE,
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=10,
            )
            response.raise_for_status()
        except requests.exceptions.RequestException as err:
            # Preserve the original exception message so callers can
            # understand why the request failed (e.g. proxy issues or
            # connection timeouts).
            raise RuntimeError(
                f"Failed to download ESRI GDP page: {err}"
            ) from err

        soup = BeautifulSoup(response.content, "html.parser")

        sections = [
            "四半期GDP成長率",
            "年次GDP成長率",
            "四半期GDP実額",
            "年次GDP実額",
        ]

        results: list[str] = []
        for section in sections:
            h3 = soup.find("h3", string=section)
            if not h3:
                continue
            block = h3.find_next("div", class_="sna_main_data_column_block")
            if not block:
                continue
            links = block.find_all("a", href=True)[:2]
            for label, link in zip(["実質", "名目"], links):
                csv_url = urljoin(self._BASE_URL, link["href"])
                try:
                    csv_response = session.get(
                        csv_url,
                        headers={"User-Agent": "Mozilla/5.0"},
                        timeout=10,
                    )
                    csv_response.raise_for_status()
                except requests.exceptions.RequestException as err:
                    raise RuntimeError(
                        f"Failed to download CSV from {csv_url}: {err}"
                    ) from err

                file_path = directory / f"{section}_{label}_{date}.csv"

                # The CSV files provided by ESRI are encoded in Shift JIS.
                # Convert the content to UTF-8 so that downstream consumers
                # can read them without handling encoding details.
                try:
                    csv_text = csv_response.content.decode("shift_jis")
                except UnicodeDecodeError as err:
                    raise RuntimeError(
                        f"Failed to decode CSV from {csv_url}: {err}"
                    ) from err

                # Remove empty lines so that downstream consumers do not
                # have to handle blank rows in the CSV output.
                cleaned_lines = [line for line in csv_text.splitlines() if line.strip()]
                cleaned_text = "\n".join(cleaned_lines) + "\n"

                file_path.write_text(cleaned_text, encoding="utf-8")
                results.append(str(file_path))

        return results

    def kp23(self, date: str) -> list[str]:
        """Download fixed capital stock PDF and Excel files.

        The PDF is converted to Markdown and its tables exported to CSV.
        From the Excel workbook, the sheets ``実質四半期ストック`` and
        ``前年同期比`` are saved as separate CSV files.

        Parameters
        ----------
        date: str
            Date string in YYYYMM format appended to each filename.

        Returns
        -------
        list[str]
            Paths to the downloaded and generated files.
        """

        pdf_dir = Path("pdf") / "kp23"
        xls_dir = Path("xls") / "kp23"
        pdf_dir.mkdir(parents=True, exist_ok=True)
        xls_dir.mkdir(parents=True, exist_ok=True)

        session = requests.Session()
        retry = Retry(
            total=5,
            backoff_factor=1,
            status_forcelist=[500, 502, 503, 504],
            allowed_methods=["GET"],
        )
        adapter = HTTPAdapter(max_retries=retry)
        session.mount("https://", adapter)
        session.mount("http://", adapter)

        try:
            response = session.get(
                self._KP23_PAGE,
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=10,
            )
            response.raise_for_status()
        except requests.exceptions.RequestException as err:
            raise RuntimeError(
                f"Failed to download ESRI kp23 page: {err}"
            ) from err

        soup = BeautifulSoup(response.content, "html.parser")

        pdf_link = None
        xls_link = None
        lists = soup.find_all("ul", class_="bulletList")
        if not lists:
            raise RuntimeError("Failed to locate kp23 links on ESRI page")

        for ul in lists:
            pdf_link = ul.find("a", href=lambda h: h and h.endswith(".pdf"))
            xls_link = ul.find("a", href=lambda h: h and h.endswith((".xls", ".xlsx")))
            if pdf_link and xls_link:
                break

        if not pdf_link or not xls_link:
            raise RuntimeError("Missing PDF or Excel link on kp23 page")

        pdf_url = urljoin(self._BASE_URL, pdf_link["href"])
        xls_url = urljoin(self._BASE_URL, xls_link["href"])

        results: list[str] = []

        pdf_path = pdf_dir / f"四半期別固定資本ストック_{date}.pdf"
        try:
            pdf_resp = session.get(
                pdf_url,
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=10,
            )
            pdf_resp.raise_for_status()
        except requests.exceptions.RequestException as err:
            raise RuntimeError(
                f"Failed to download PDF from {pdf_url}: {err}"
            ) from err
        pdf_path.write_bytes(pdf_resp.content)
        results.append(str(pdf_path))

        helper = meti()
        results.append(helper.pdf_to_markdown(str(pdf_path)))
        results.extend(helper.pdf_tables_to_csv(str(pdf_path)))

        xls_path = xls_dir / f"四半期別固定資本ストック_{date}.xlsx"
        try:
            xls_resp = session.get(
                xls_url,
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=10,
            )
            xls_resp.raise_for_status()
        except requests.exceptions.RequestException as err:
            raise RuntimeError(
                f"Failed to download Excel from {xls_url}: {err}"
            ) from err
        xls_path.write_bytes(xls_resp.content)
        results.append(str(xls_path))

        wb = load_workbook(xls_path, data_only=True, read_only=True)
        sheets = {
            "実質四半期ストック": xls_dir / f"四半期別固定資本ストック_実質_{date}.csv",
            "前年同期比": xls_dir / f"四半期別固定資本ストック_同期比_{date}.csv",
        }
        for sheet_name, csv_path in sheets.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            with csv_path.open("w", newline="", encoding="utf-8") as csvfile:
                writer = csv.writer(csvfile)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow([cell if cell is not None else "" for cell in row])
            results.append(str(csv_path))
        wb.close()

        return results

if __name__ == "__main__":
    from datetime import datetime
    import sys

    today = datetime.today()

    scraper = esri()
    try:
        gdp_paths = scraper.gdp(date=today.strftime("%Y%m"))
        for path in gdp_paths:
            print(path)
    except RuntimeError as err:
        print(err)
        sys.exit(1)
